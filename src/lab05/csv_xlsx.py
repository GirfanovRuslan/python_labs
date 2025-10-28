import csv
import openpyxl
import os
from pathlib import Path

def validate_file_exists(file_path: str) -> Path:
    """Проверяет существование файла и возвращает Path объект"""
    if not file_path or not isinstance(file_path, str):
        raise ValueError("Путь к файлу должен быть непустой строкой")
    
    path = Path(file_path)
    
    if not path.exists():
        raise FileNotFoundError(f"Файл {file_path} не найден")
    
    if not path.is_file():
        raise ValueError(f"{file_path} не является файлом")
    
    return path

def validate_file_not_empty(path: Path) -> None:
    """Проверяет что файл не пустой"""
    try:
        if path.stat().st_size == 0:
            raise ValueError(f"Файл {path} пустой")
    except PermissionError:
        raise PermissionError(f"Нет доступа к файлу {path}")
    except OSError as e:
        raise OSError(f"Ошибка доступа к файлу {path}: {e}")

def validate_input_extension(file_path: str, expected_ext: str) -> None:
    """Проверяет расширение входного файла"""
    if not expected_ext.startswith('.'):
        expected_ext = '.' + expected_ext
    
    path = validate_file_exists(file_path)
    validate_file_not_empty(path)
    
    ext = path.suffix.lower()
    expected_ext = expected_ext.lower()
    
    if ext != expected_ext:
        raise ValueError(f"Ошибка: входной файл {file_path} должен иметь расширение {expected_ext}")

def validate_output_extension(file_path: str, expected_ext: str) -> None:
    """Проверяет расширение выходного файла"""
    if not expected_ext.startswith('.'):
        expected_ext = '.' + expected_ext
    
    if not file_path or not isinstance(file_path, str):
        raise ValueError("Путь к файлу должен быть непустой строкой")
    
    path = Path(file_path)
    ext = path.suffix.lower()
    expected_ext = expected_ext.lower()
    
    if ext != expected_ext:
        raise ValueError(f"Ошибка: выходной файл {file_path} должен иметь расширение {expected_ext}")
    
    # Создаем директорию если её нет
    path.parent.mkdir(parents=True, exist_ok=True)

def validate_csv_structure(csv_file: str, encoding: str = 'utf-8') -> tuple[list, list]:
    """Проверяет структуру CSV файла и возвращает данные и заголовки"""
    data = []
    try:
        with open(csv_file, 'r', encoding=encoding) as f:
            # Проверяем, что файл не пустой
            first_char = f.read(1)
            if not first_char:
                raise ValueError("CSV файл пустой")
            f.seek(0)
            
            # Определяем диалект CSV
            sample = f.read(1024)
            f.seek(0)
            
            try:
                sniffer = csv.Sniffer()
                dialect = sniffer.sniff(sample)
                has_header = sniffer.has_header(sample)
            except csv.Error:
                dialect = csv.excel
                has_header = True
            
            if not has_header:
                raise ValueError("CSV файл должен содержать заголовок")
            
            reader = csv.reader(f, dialect=dialect)
            headers = next(reader)
            
            # Проверяем заголовки
            if not headers or all(not header.strip() for header in headers):
                raise ValueError("CSV файл содержит пустые заголовки")
            
            # Читаем и проверяем данные
            for row_idx, row in enumerate(reader, 2):  # начинаем с 2 строки (после заголовка)
                if len(row) != len(headers):
                    raise ValueError(f"Строка {row_idx} имеет {len(row)} полей, ожидается {len(headers)}")
                data.append(row)
                
    except UnicodeDecodeError:
        raise ValueError(f"Неверная кодировка файла {csv_file}. Попробуйте другую кодировку.")
    
    return data, headers

def csv_to_xlsx(csv_file: str, xlsx_file: str, encoding: str = 'utf-8') -> None:
    """Конвертирует CSV в XLSX используя openpyxl"""
    validate_input_extension(csv_file, '.csv')
    validate_output_extension(xlsx_file, '.xlsx')
    
    try:
        # Проверяем структуру CSV и получаем данные
        data, headers = validate_csv_structure(csv_file, encoding)
        
        # Создаем новую книгу Excel
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        
        # Записываем заголовки
        for col_idx, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_idx, value=header)
        
        # Записываем данные
        for row_idx, row in enumerate(data, 2):  # начинаем со 2 строки
            for col_idx, value in enumerate(row, 1):
                # Пытаемся преобразовать числа
                try:
                    if value.replace('.', '').replace('-', '').isdigit():
                        value = float(value) if '.' in value else int(value)
                except (ValueError, AttributeError):
                    pass  # Оставляем как строку если не число
                
                sheet.cell(row=row_idx, column=col_idx, value=value)
        
        # Автонастройка ширины колонок
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Максимум 50 символов
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Сохраняем файл
        workbook.save(xlsx_file)
        
    except Exception as e:
        # Удаляем частично записанный файл при ошибке
        if os.path.exists(xlsx_file):
            try:
                os.remove(xlsx_file)
            except:
                pass  # Игнорируем ошибки удаления
        raise Exception(f"Ошибка конвертации CSV в XLSX: {e}")
    
    finally:
        # Закрываем workbook если он был создан
        if 'workbook' in locals():
            try:
                workbook.close()
            except:
                pass